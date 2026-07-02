"""Full-database Excel backup / restore for the leads app.

Produces a single multi-sheet ``.xlsx`` workbook that captures every lead (with
its folder membership), all lead groups, WhatsApp chat history, conversation
logs, the WhatsApp campaign config, and per-folder script templates. The same
workbook can be re-imported after a redeploy onto a fresh database.

Import is *skip-existing*: leads are matched on ``(name, address)``; rows that
already exist are left untouched, and only their related history is skipped so
the operation is safe to re-run.
"""

from __future__ import annotations

import json
from datetime import date, datetime, time
from typing import Iterator, Optional

from django.db import transaction
from django.utils import timezone as django_timezone
from django.utils.dateparse import parse_date, parse_datetime, parse_time

from leads.models import (
    ChatMessage,
    Lead,
    LeadConversationLog,
    LeadGroup,
    WhatsAppConfig,
    WhatsAppScriptTemplate,
)
from leads.pipeline import ensure_pipeline_system_groups

LEAD_HEADERS = [
    "id",
    "name",
    "phone_number",
    "phone_numbers",
    "address",
    "website",
    "shop_keyword",
    "category",
    "source_url",
    "is_processed",
    "is_chain",
    "chain_detected_internal",
    "chain_detected_ai",
    "location_count_estimate",
    "is_very_important",
    "whatsapp_draft",
    "whatsapp_status",
    "whatsapp_sent_at",
    "whatsapp_instance_id",
    "whatsapp_last_error",
    "search_city",
    "search_state",
    "search_query",
    "search_country",
    "group",
    "display_order",
    "created_at",
]

CHAT_HEADERS = [
    "lead_id",
    "body",
    "is_outbound",
    "template_name",
    "meta_message_id",
    "created_at",
]

LOG_HEADERS = [
    "lead_id",
    "conversation_date",
    "remarks",
    "created_at",
]

GROUP_HEADERS = ["name", "sort_order"]

CONFIG_HEADERS = [
    "allowed_days",
    "window1_start",
    "window1_end",
    "window2_start",
    "window2_end",
    "is_paused",
    "meta_message_templates",
    "meta_templates_synced_at",
    "outbound_template_name",
]

SCRIPT_HEADERS = ["group_name", "template_text"]


# --------------------------------------------------------------------------- #
# Serialization helpers (model value -> cell)
# --------------------------------------------------------------------------- #
def _dump_dt(value) -> str:
    return value.isoformat() if value else ""


def _dump_date(value) -> str:
    return value.isoformat() if value else ""


def _dump_time(value) -> str:
    return value.strftime("%H:%M:%S") if value else ""


def _dump_json(value) -> str:
    if value in (None, ""):
        return ""
    return json.dumps(value, ensure_ascii=False)


# --------------------------------------------------------------------------- #
# Parsing helpers (cell -> model value)
# --------------------------------------------------------------------------- #
def _p_str(value) -> str:
    return "" if value is None else str(value)


def _p_int(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _p_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "t"}


def _make_aware(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if django_timezone.is_naive(dt) and django_timezone.get_current_timezone():
        try:
            return django_timezone.make_aware(dt)
        except Exception:
            return dt
    return dt


def _p_dt(value) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return _make_aware(value)
    parsed = parse_datetime(str(value))
    return _make_aware(parsed) if parsed else None


def _p_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(str(value))


def _p_time(value) -> Optional[time]:
    if value in (None, ""):
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return parse_time(str(value))


def _p_json(value, default):
    if value in (None, ""):
        return default
    if isinstance(value, (list, dict)):
        return value
    try:
        return json.loads(str(value))
    except (TypeError, ValueError):
        return default


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def build_backup_workbook(lead_ids: list[int] | None = None):
    """Return an openpyxl Workbook containing the leads backup.

    When ``lead_ids`` is provided, only those leads (and their chats/logs/groups) are
    included. Otherwise every lead in the database is exported.
    """
    from openpyxl import Workbook  # local import: optional dependency

    wb = Workbook()
    id_filter = [int(x) for x in (lead_ids or []) if int(x) > 0] if lead_ids else None

    lead_qs = Lead.objects.select_related("group").order_by("display_order", "-created_at")
    if id_filter:
        lead_qs = lead_qs.filter(pk__in=id_filter)

    ws_leads = wb.active
    ws_leads.title = "Leads"
    ws_leads.append(LEAD_HEADERS)
    for c in lead_qs.iterator(chunk_size=500):
        ws_leads.append(
            [
                c.pk,
                c.name,
                c.phone_number or "",
                _dump_json(c.phone_numbers),
                c.address or "",
                c.website or "",
                c.shop_keyword or "",
                c.category,
                c.source_url or "",
                bool(c.is_processed),
                bool(c.is_chain),
                bool(c.chain_detected_internal),
                bool(c.chain_detected_ai),
                c.location_count_estimate,
                bool(c.is_very_important),
                (c.whatsapp_draft or "")[:32000],
                c.whatsapp_status,
                _dump_dt(c.whatsapp_sent_at),
                c.whatsapp_instance_id or "",
                (c.whatsapp_last_error or "")[:32000],
                c.search_city or "",
                c.search_state or "",
                c.search_query or "",
                c.search_country or "",
                c.group.name if c.group_id else "",
                c.display_order,
                _dump_dt(c.created_at),
            ]
        )

    ws_groups = wb.create_sheet("Groups")
    ws_groups.append(GROUP_HEADERS)
    if id_filter:
        group_ids = [
            gid
            for gid in lead_qs.values_list("group_id", flat=True).distinct()
            if gid is not None
        ]
        groups_qs = LeadGroup.objects.filter(pk__in=group_ids).order_by("sort_order", "name")
    else:
        groups_qs = LeadGroup.objects.all().order_by("sort_order", "name")
    for g in groups_qs:
        ws_groups.append([g.name, g.sort_order])

    ws_chats = wb.create_sheet("ChatMessages")
    ws_chats.append(CHAT_HEADERS)
    chat_qs = ChatMessage.objects.all().order_by("lead_id", "created_at", "id")
    if id_filter:
        chat_qs = chat_qs.filter(lead_id__in=id_filter)
    for m in chat_qs.iterator(chunk_size=1000):
        ws_chats.append(
            [
                m.lead_id,
                (m.body or "")[:32000],
                bool(m.is_outbound),
                m.template_name or "",
                m.meta_message_id or "",
                _dump_dt(m.created_at),
            ]
        )

    ws_logs = wb.create_sheet("ConversationLogs")
    ws_logs.append(LOG_HEADERS)
    log_qs = LeadConversationLog.objects.all().order_by("lead_id", "created_at", "id")
    if id_filter:
        log_qs = log_qs.filter(lead_id__in=id_filter)
    for log in log_qs.iterator(chunk_size=1000):
        ws_logs.append(
            [
                log.lead_id,
                _dump_date(log.conversation_date),
                (log.remarks or "")[:32000],
                _dump_dt(log.created_at),
            ]
        )

    ws_cfg = wb.create_sheet("WhatsAppConfig")
    ws_cfg.append(CONFIG_HEADERS)
    cfg = WhatsAppConfig.objects.first()
    if cfg:
        ws_cfg.append(
            [
                _dump_json(cfg.allowed_days),
                _dump_time(cfg.window1_start),
                _dump_time(cfg.window1_end),
                _dump_time(cfg.window2_start),
                _dump_time(cfg.window2_end),
                bool(cfg.is_paused),
                _dump_json(cfg.meta_message_templates),
                _dump_dt(cfg.meta_templates_synced_at),
                cfg.outbound_template_name or "",
            ]
        )

    ws_scripts = wb.create_sheet("ScriptTemplates")
    ws_scripts.append(SCRIPT_HEADERS)
    for tpl in WhatsAppScriptTemplate.objects.all().order_by("group_name"):
        ws_scripts.append([tpl.group_name, tpl.template_text or ""])

    return wb


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
def _iter_sheet_dicts(ws) -> Iterator[dict]:
    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None)
    if not raw_headers:
        return
    headers = [_p_str(h).strip() for h in raw_headers]
    for row in rows:
        if row is None:
            continue
        if all(cell is None or cell == "" for cell in row):
            continue
        yield dict(zip(headers, row))


def _sheet(wb, name):
    return wb[name] if name in wb.sheetnames else None


@transaction.atomic
def restore_from_workbook(file_obj) -> dict:
    """Restore a backup workbook. Existing leads (by name+address) are skipped.

    Returns a summary dict of created / skipped counts per entity.
    """
    from openpyxl import load_workbook  # local import: optional dependency

    wb = load_workbook(file_obj, read_only=True, data_only=True)

    summary = {
        "groups_created": 0,
        "leads_created": 0,
        "leads_skipped": 0,
        "chats_created": 0,
        "logs_created": 0,
        "config_created": 0,
        "scripts_created": 0,
    }

    from leads.category_types import UNKNOWN_SLUG, lead_category_choices

    valid_categories = {c[0] for c in lead_category_choices()}
    valid_statuses = {s[0] for s in Lead.WhatsappStatus.choices}

    # Groups first so lead FKs can be linked by name.
    group_map: dict[str, LeadGroup] = {}
    ws_groups = _sheet(wb, "Groups")
    if ws_groups is not None:
        for row in _iter_sheet_dicts(ws_groups):
            name = _p_str(row.get("name")).strip()
            if not name:
                continue
            grp, created = LeadGroup.objects.get_or_create(
                name=name,
                defaults={"sort_order": _p_int(row.get("sort_order")) or 0},
            )
            group_map[name] = grp
            if created:
                summary["groups_created"] += 1
    ensure_pipeline_system_groups()

    def resolve_group(name: str) -> Optional[LeadGroup]:
        name = (name or "").strip()
        if not name:
            return None
        if name in group_map:
            return group_map[name]
        grp, _created = LeadGroup.objects.get_or_create(name=name)
        group_map[name] = grp
        return grp

    # Leads.
    lead_map: dict[int, Lead] = {}
    created_lead_ids: set[int] = set()
    ws_leads = _sheet(wb, "Leads")
    if ws_leads is not None:
        for row in _iter_sheet_dicts(ws_leads):
            name = _p_str(row.get("name")).strip()
            if not name:
                continue
            address = _p_str(row.get("address"))
            category = _p_str(row.get("category")).strip().lower()
            if category not in valid_categories:
                category = UNKNOWN_SLUG
            status = _p_str(row.get("whatsapp_status")).strip().lower()
            if status not in valid_statuses:
                status = Lead.WhatsappStatus.IDLE
            sc = _p_str(row.get("search_city")).strip()
            ss = _p_str(row.get("search_state")).strip()
            sq = _p_str(row.get("search_query")).strip()
            sco = _p_str(row.get("search_country")).strip()
            defaults = {
                "phone_number": _p_str(row.get("phone_number"))[:64],
                "phone_numbers": _p_json(row.get("phone_numbers"), []),
                "website": _p_str(row.get("website"))[:500],
                "shop_keyword": _p_str(row.get("shop_keyword"))[:160],
                "category": category,
                "source_url": _p_str(row.get("source_url")),
                "is_processed": _p_bool(row.get("is_processed")),
                "is_chain": _p_bool(row.get("is_chain")),
                "chain_detected_internal": _p_bool(row.get("chain_detected_internal")),
                "chain_detected_ai": _p_bool(row.get("chain_detected_ai")),
                "location_count_estimate": _p_int(row.get("location_count_estimate")),
                "is_very_important": _p_bool(row.get("is_very_important")),
                "whatsapp_draft": _p_str(row.get("whatsapp_draft")),
                "whatsapp_status": status,
                "whatsapp_sent_at": _p_dt(row.get("whatsapp_sent_at")),
                "whatsapp_instance_id": _p_str(row.get("whatsapp_instance_id"))[:120],
                "whatsapp_last_error": _p_str(row.get("whatsapp_last_error")),
                "search_city": sc or None,
                "search_state": ss or None,
                "search_query": sq or None,
                "search_country": sco or None,
                "group": resolve_group(_p_str(row.get("group"))),
                "display_order": _p_int(row.get("display_order")) or 0,
            }
            lead, created = Lead.objects.get_or_create(
                name=name[:255],
                address=address,
                defaults=defaults,
            )
            old_id = _p_int(row.get("id"))
            if old_id is not None:
                lead_map[old_id] = lead
            if created:
                summary["leads_created"] += 1
                if old_id is not None:
                    created_lead_ids.add(old_id)
                created_at = _p_dt(row.get("created_at"))
                if created_at:
                    Lead.objects.filter(pk=lead.pk).update(created_at=created_at)
            else:
                summary["leads_skipped"] += 1

    # Chat messages — only attach to leads we just created (skip-existing safety).
    ws_chats = _sheet(wb, "ChatMessages")
    if ws_chats is not None:
        for row in _iter_sheet_dicts(ws_chats):
            old_lead = _p_int(row.get("lead_id"))
            if old_lead is None or old_lead not in created_lead_ids:
                continue
            lead = lead_map.get(old_lead)
            if lead is None:
                continue
            msg = ChatMessage.objects.create(
                lead=lead,
                body=_p_str(row.get("body")),
                is_outbound=_p_bool(row.get("is_outbound")),
                template_name=_p_str(row.get("template_name"))[:64],
                meta_message_id=_p_str(row.get("meta_message_id"))[:128],
            )
            created_at = _p_dt(row.get("created_at"))
            if created_at:
                ChatMessage.objects.filter(pk=msg.pk).update(created_at=created_at)
            summary["chats_created"] += 1

    # Conversation logs — same skip-existing rule.
    ws_logs = _sheet(wb, "ConversationLogs")
    if ws_logs is not None:
        for row in _iter_sheet_dicts(ws_logs):
            old_lead = _p_int(row.get("lead_id"))
            if old_lead is None or old_lead not in created_lead_ids:
                continue
            lead = lead_map.get(old_lead)
            if lead is None:
                continue
            conv_date = _p_date(row.get("conversation_date"))
            if conv_date is None:
                continue
            log = LeadConversationLog.objects.create(
                lead=lead,
                conversation_date=conv_date,
                remarks=_p_str(row.get("remarks")),
            )
            created_at = _p_dt(row.get("created_at"))
            if created_at:
                LeadConversationLog.objects.filter(pk=log.pk).update(
                    created_at=created_at
                )
            summary["logs_created"] += 1

    # WhatsApp config — only when none exists yet.
    ws_cfg = _sheet(wb, "WhatsAppConfig")
    if ws_cfg is not None and not WhatsAppConfig.objects.exists():
        for row in _iter_sheet_dicts(ws_cfg):
            WhatsAppConfig.objects.create(
                id=WhatsAppConfig.SINGLETON_ID,
                allowed_days=_p_json(
                    row.get("allowed_days"), list(WhatsAppConfig.DEFAULT_ALLOWED_DAYS)
                ),
                window1_start=_p_time(row.get("window1_start")) or time(8, 0),
                window1_end=_p_time(row.get("window1_end")) or time(13, 0),
                window2_start=_p_time(row.get("window2_start")) or time(15, 0),
                window2_end=_p_time(row.get("window2_end")) or time(20, 0),
                is_paused=_p_bool(row.get("is_paused")),
                meta_message_templates=_p_json(row.get("meta_message_templates"), []),
                meta_templates_synced_at=_p_dt(row.get("meta_templates_synced_at")),
                outbound_template_name=_p_str(row.get("outbound_template_name"))[:64]
                or "just_to_say_hi",
            )
            summary["config_created"] += 1
            break  # singleton: only the first row matters

    # Script templates — keyed by unique group_name; skip existing.
    ws_scripts = _sheet(wb, "ScriptTemplates")
    if ws_scripts is not None:
        for row in _iter_sheet_dicts(ws_scripts):
            group_name = _p_str(row.get("group_name")).strip()
            if not group_name:
                continue
            _tpl, created = WhatsAppScriptTemplate.objects.get_or_create(
                group_name=group_name[:100],
                defaults={"template_text": _p_str(row.get("template_text"))},
            )
            if created:
                summary["scripts_created"] += 1

    return summary
