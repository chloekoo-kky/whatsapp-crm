import json
from datetime import date
from unittest.mock import Mock, patch

from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.db.models import Exists, OuterRef

from leads.models import CategoryRule, ChatMessage, Lead, LeadCategoryType, LeadConversationLog, LeadGroup, WhatsAppScriptTemplate
from leads.chat_messages import record_inbound_chat_message, record_outbound_chat_message
from leads.display import (
    lead_google_maps_url,
    lead_phone_list,
    lead_whatsapp_active_chat,
    lead_whatsapp_dispatched,
    normalize_manual_phone,
)
from leads.views import _leads_qs_for_tab, _leads_tab_base_qs
from leads.whatsapp_service import compose_outbound_message, render_script_template
from leads.whatsapp_webhook import parse_meta_cloud_webhook
from leads.pipeline import (
    QUEUE_GROUP_NAME,
    TRASH_GROUP_NAME,
    UNCATEGORIZED_GROUP_NAME,
    WHATSAPP_CHATS_GROUP_NAME,
    apply_group_assignment_side_effects,
    ensure_pipeline_system_groups,
    enqueue_leads_for_whatsapp,
    get_or_create_uncategorized_group,
    phone_exists_in_database,
)


class PipelineGroupTests(TestCase):
    def test_system_groups_are_created(self):
        groups = ensure_pipeline_system_groups()
        self.assertEqual(groups["uncategorized"].name, UNCATEGORIZED_GROUP_NAME)
        self.assertEqual(groups["queue"].name, QUEUE_GROUP_NAME)
        self.assertEqual(groups["whatsapp_chats"].name, WHATSAPP_CHATS_GROUP_NAME)
        self.assertEqual(groups["trash"].name, TRASH_GROUP_NAME)

    def test_new_lead_defaults_to_idle_not_pending(self):
        lead = Lead.objects.create(
            name="Gamma Clinic",
            address="3 Main St",
            group=get_or_create_uncategorized_group(),
        )
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.IDLE)

    def test_enqueue_sets_pending_without_moving_group(self):
        groups = ensure_pipeline_system_groups()
        uncategorized = groups["uncategorized"]
        lead = Lead.objects.create(
            name="Alpha Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=uncategorized,
            whatsapp_status=Lead.WhatsappStatus.IDLE,
            display_order=1,
        )
        updated = enqueue_leads_for_whatsapp([lead.pk])
        lead.refresh_from_db()
        self.assertEqual(updated, 1)
        self.assertEqual(lead.group_id, uncategorized.pk)
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.PENDING)
        self.assertEqual(lead.display_order, 1)

    def test_dequeue_reverts_pending_lead_to_idle(self):
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Delta Clinic",
            address="4 Main St",
            phone_number="+60111222333",
            phone_numbers=["+60111222333"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.PENDING,
        )
        client = Client(enforce_csrf_checks=True)
        client.get("/")
        response = client.post(
            f"/leads/ajax/lead/{lead.pk}/dequeue/",
            HTTP_HX_REQUEST="true",
            HTTP_X_CSRFTOKEN=client.cookies["csrftoken"].value,
        )
        self.assertEqual(response.status_code, 200)
        lead.refresh_from_db()
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.IDLE)
        self.assertIn("lead-join-queue-btn", response.content.decode())

    def test_leaving_queue_folder_returns_lead_to_idle(self):
        groups = ensure_pipeline_system_groups()
        queue = groups["queue"]
        uncategorized = groups["uncategorized"]
        lead = Lead.objects.create(
            name="Beta Clinic",
            address="2 Main St",
            group=queue,
            phone_number="+60198765432",
            phone_numbers=["+60198765432"],
            whatsapp_status=Lead.WhatsappStatus.PENDING,
        )
        apply_group_assignment_side_effects([lead], uncategorized)
        lead.refresh_from_db()
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.IDLE)

    def test_enqueue_after_sent_allows_re_promotion(self):
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Sent Clinic",
            address="5 Main St",
            phone_number="+60122334455",
            phone_numbers=["+60122334455"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        updated = enqueue_leads_for_whatsapp([lead.pk])
        lead.refresh_from_db()
        self.assertEqual(updated, 1)
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.PENDING)


class LeadDisplayPipelineTests(TestCase):
    def test_dispatched_when_sent_without_client_reply(self):
        lead = Lead.objects.create(
            name="Outbound Only",
            address="1 Road",
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        self.assertTrue(lead_whatsapp_dispatched(lead))
        self.assertFalse(lead_whatsapp_active_chat(lead))

    def test_dispatched_persists_when_re_enqueued(self):
        from django.utils import timezone

        lead = Lead.objects.create(
            name="Re-promo Clinic",
            address="4 Road",
            whatsapp_status=Lead.WhatsappStatus.PENDING,
            whatsapp_sent_at=timezone.now(),
        )
        self.assertTrue(lead_whatsapp_dispatched(lead))
        self.assertFalse(lead_whatsapp_active_chat(lead))

    def test_active_chat_requires_latest_inbound_message(self):
        from django.utils import timezone

        lead = Lead.objects.create(
            name="Replied Clinic",
            address="2 Road",
            whatsapp_status=Lead.WhatsappStatus.SENT,
            whatsapp_sent_at=timezone.now(),
        )
        record_outbound_chat_message(lead, body="Hello from CRM")
        record_inbound_chat_message(lead, body="Yes please")
        annotated = _leads_tab_base_qs().get(pk=lead.pk)
        self.assertTrue(lead_whatsapp_active_chat(annotated))

    def test_active_chat_cleared_after_staff_reply(self):
        from django.utils import timezone

        lead = Lead.objects.create(
            name="Handled Clinic",
            address="6 Road",
            whatsapp_status=Lead.WhatsappStatus.SENT,
            whatsapp_sent_at=timezone.now(),
        )
        record_outbound_chat_message(lead, body="Hello from CRM")
        record_inbound_chat_message(lead, body="Interested")
        record_outbound_chat_message(lead, body="Great, let's talk")
        annotated = _leads_tab_base_qs().get(pk=lead.pk)
        self.assertFalse(lead_whatsapp_active_chat(annotated))

    def test_outbound_only_thread_has_no_active_chat_pulse(self):
        from django.utils import timezone

        lead = Lead.objects.create(
            name="Outbound Thread",
            address="7 Road",
            whatsapp_status=Lead.WhatsappStatus.SENT,
            whatsapp_sent_at=timezone.now(),
        )
        record_outbound_chat_message(lead, body="Hello from CRM")
        annotated = _leads_tab_base_qs().get(pk=lead.pk)
        self.assertFalse(lead_whatsapp_active_chat(annotated))

    def test_human_log_without_client_reply_is_not_active_chat(self):
        lead = Lead.objects.create(
            name="Staff Note",
            address="3 Road",
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        LeadConversationLog.objects.create(
            lead=lead,
            conversation_date=date.today(),
            remarks="[WhatsApp · staff] Follow up tomorrow",
        )
        annotated = Lead.objects.annotate(
            has_client_conversation_log=Exists(
                LeadConversationLog.objects.filter(
                    lead_id=OuterRef("pk"),
                    remarks__icontains="[WhatsApp · client]",
                )
            ),
            has_inbound_chat_message=Exists(
                ChatMessage.objects.filter(lead_id=OuterRef("pk"), is_outbound=False)
            ),
            has_human_conversation_log=Exists(
                LeadConversationLog.objects.filter(lead_id=OuterRef("pk")).exclude(
                    remarks__icontains="Touchpoint Automator"
                )
            ),
        ).get(pk=lead.pk)
        self.assertFalse(lead_whatsapp_active_chat(annotated))

    @override_settings(WHATSAPP_CAMPAIGN_TIMEZONE="Asia/Kuala_Lumpur")
    def test_campaign_datetime_filter_uses_local_timezone(self):
        from datetime import datetime, timezone as dt_timezone

        from leads.templatetags.clinic_display import campaign_datetime

        utc = datetime(2026, 7, 1, 2, 41, tzinfo=dt_timezone.utc)
        self.assertEqual(campaign_datetime(utc), "Jul 1, 2026 · 10:41 AM")

    def test_google_maps_url_uses_name_instead_of_coordinate_source(self):
        lead = Lead.objects.create(
            name="U.n.i Klinik Iskandar Puteri",
            address="97 Jalan Suria 2, Iskandar Puteri, Johor",
            source_url="https://www.google.com/maps/search/?api=1&query=1.453703,103.599190",
        )
        url = lead_google_maps_url(lead)
        self.assertIn("query=", url)
        self.assertNotIn("1.453703", url)
        self.assertIn("U.n.i", url)
        self.assertIn("Iskandar", url)

    def test_normalize_manual_phone_malaysian_landlines(self):
        self.assertEqual(normalize_manual_phone("07-585 4964"), "+6075854964")
        self.assertEqual(normalize_manual_phone("03-1234 5678"), "+60312345678")
        self.assertEqual(normalize_manual_phone("6075854964"), "+6075854964")
        self.assertEqual(normalize_manual_phone("+6075854964"), "+6075854964")

    def test_normalize_manual_phone_repairs_double_country_prefix(self):
        self.assertEqual(normalize_manual_phone("+606075854964"), "+6075854964")
        self.assertEqual(normalize_manual_phone("606075854964"), "+6075854964")

    def test_lead_phone_list_repairs_stored_double_prefix(self):
        lead = Lead.objects.create(
            name="Johor Clinic",
            address="1 Road",
            phone_number="+606075854964",
            phone_numbers=["+606075854964"],
        )
        self.assertEqual(lead_phone_list(lead), ["+6075854964"])


class ActiveChatTabTests(TestCase):
    def test_active_chat_tab_lists_awaiting_leads_without_moving_group(self):
        from django.utils import timezone

        groups = ensure_pipeline_system_groups()
        quality = LeadGroup.objects.create(name="Quality Leads", sort_order=20)
        lead = Lead.objects.create(
            name="Tab Clinic",
            address="8 Road",
            phone_number="+60199887766",
            phone_numbers=["+60199887766"],
            group=quality,
            whatsapp_status=Lead.WhatsappStatus.SENT,
            whatsapp_sent_at=timezone.now(),
        )
        record_outbound_chat_message(lead, body="Hello")
        record_inbound_chat_message(lead, body="Please call me")

        active_chat_qs = _leads_qs_for_tab(str(groups["whatsapp_chats"].pk), None)
        self.assertEqual(list(active_chat_qs.values_list("pk", flat=True)), [lead.pk])
        lead.refresh_from_db()
        self.assertEqual(lead.group_id, quality.pk)


class WhatsAppScriptTemplateTests(TestCase):
    def test_render_script_template_substitutes_placeholders(self):
        lead = Lead.objects.create(
            name="Glow Clinic",
            address="12 Jalan Ampang",
            search_city="Kuala Lumpur",
        )
        text = render_script_template(
            "Hello {{ name }} from {{ area }}!",
            lead,
        )
        self.assertEqual(text, "Hello Glow Clinic from Kuala Lumpur!")

    def test_compose_outbound_message_uses_group_template(self):
        aesthetic = LeadGroup.objects.create(name="Aesthetic", sort_order=10)
        WhatsAppScriptTemplate.objects.create(
            group_name="Aesthetic",
            template_text="Hi {{ name }}, welcome to {{ area }}.",
        )
        lead = Lead.objects.create(
            name="Skin Lab",
            address="1 Main St",
            search_city="Petaling Jaya",
            group=aesthetic,
        )
        body = compose_outbound_message(lead)
        self.assertEqual(body, "Hi Skin Lab, welcome to Petaling Jaya.")

    def test_compose_outbound_message_falls_back_to_default(self):
        lead = Lead.objects.create(
            name="Unknown Shop",
            address="9 Road",
            search_city="Johor Bahru",
            group=ensure_pipeline_system_groups()["uncategorized"],
        )
        body = compose_outbound_message(lead)
        self.assertIn("Unknown Shop", body)
        self.assertIn("Johor Bahru", body)


class WhatsAppWebhookTests(TestCase):
    def _meta_inbound_payload(self, **message_overrides):
        message = {
            "from": "60123456789",
            "id": "wamid.MSG123",
            "timestamp": "1709550600",
            "type": "text",
            "text": {"body": "Yes, interested"},
        }
        message.update(message_overrides)
        return {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "id": "WABA_ID",
                    "changes": [
                        {
                            "field": "messages",
                            "value": {
                                "messaging_product": "whatsapp",
                                "messages": [message],
                            },
                        }
                    ],
                }
            ],
        }

    def test_parse_meta_cloud_webhook_extracts_client_text(self):
        parsed = parse_meta_cloud_webhook(self._meta_inbound_payload())
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0].remote_phone, "+60123456789")
        self.assertEqual(parsed[0].text_body, "Yes, interested")
        self.assertFalse(parsed[0].from_me)

    def _meta_echo_payload(self, **echo_overrides):
        echo = {
            "from": "60126336429",
            "to": "60123456789",
            "id": "wamid.ECHO123",
            "timestamp": "1709550700",
            "type": "text",
            "text": {"body": "Thanks, we can help with that."},
        }
        echo.update(echo_overrides)
        return {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "id": "WABA_ID",
                    "changes": [
                        {
                            "field": "smb_message_echoes",
                            "value": {
                                "messaging_product": "whatsapp",
                                "metadata": {
                                    "display_phone_number": "60126336429",
                                    "phone_number_id": "999888777",
                                },
                                "message_echoes": [echo],
                            },
                        }
                    ],
                }
            ],
        }

    def test_parse_meta_cloud_webhook_extracts_smb_message_echoes(self):
        parsed = parse_meta_cloud_webhook(self._meta_echo_payload())
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0].remote_phone, "+60123456789")
        self.assertEqual(parsed[0].text_body, "Thanks, we can help with that.")
        self.assertTrue(parsed[0].from_me)

    def test_webhook_syncs_smb_message_echo_to_outbound_chat(self):
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Echo Clinic",
            address="3 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        client = Client()
        response = client.post(
            "/webhook/whatsapp/",
            data=json.dumps(self._meta_echo_payload()),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["synced"], 1)

        chat = ChatMessage.objects.get(lead=lead, is_outbound=True)
        self.assertEqual(chat.body, "Thanks, we can help with that.")
        self.assertEqual(chat.meta_message_id, "wamid.ECHO123")
        log = LeadConversationLog.objects.get(lead=lead)
        self.assertIn("[WhatsApp · agent]", log.remarks)

    def test_webhook_syncs_log_and_keeps_lead_in_origin_group(self):
        groups = ensure_pipeline_system_groups()
        quality = LeadGroup.objects.create(name="Quality Leads", sort_order=20)
        lead = Lead.objects.create(
            name="Webhook Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=quality,
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        client = Client()
        response = client.post(
            "/webhook/whatsapp/",
            data=json.dumps(self._meta_inbound_payload()),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "success")
        self.assertEqual(response.json()["synced"], 1)

        lead.refresh_from_db()
        self.assertEqual(lead.group_id, quality.pk)
        log = LeadConversationLog.objects.get(lead=lead)
        self.assertIn("[WhatsApp · client]", log.remarks)
        self.assertIn("Yes, interested", log.remarks)
        chat = ChatMessage.objects.get(lead=lead, is_outbound=False)
        self.assertEqual(chat.body, "Yes, interested")
        self.assertEqual(chat.meta_message_id, "wamid.MSG123")

    def test_webhook_verify_get_returns_challenge(self):
        client = Client()
        with override_settings(WHATSAPP_WEBHOOK_VERIFY_TOKEN="verify-me"):
            response = client.get(
                "/webhook/whatsapp/",
                {
                    "hub.mode": "subscribe",
                    "hub.verify_token": "verify-me",
                    "hub.challenge": "1234567890",
                },
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "1234567890")

    @override_settings(WHATSAPP_WEBHOOK_VERIFY_TOKEN="CLINIC_CRM_WEBHOOK_73R469Mf")
    def test_webhook_receiver_verify_get_returns_challenge(self):
        client = Client()
        response = client.get(
            "/whatsapp/webhook/",
            {
                "hub.mode": "subscribe",
                "hub.verify_token": "CLINIC_CRM_WEBHOOK_73R469Mf",
                "hub.challenge": "99887766",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "99887766")

    @override_settings(WHATSAPP_APP_SECRET="meta-app-secret")
    def test_webhook_rejects_unsigned_from_public_ip(self):
        client = Client()
        response = client.post(
            "/webhook/whatsapp/",
            data=json.dumps(self._meta_inbound_payload()),
            content_type="application/json",
            REMOTE_ADDR="8.8.8.8",
        )
        self.assertEqual(response.status_code, 403)

    def test_webhook_accepts_meta_forwarded_ip_when_peer_is_local(self):
        """ngrok forwards Meta's public IP in X-Forwarded-For; auth uses REMOTE_ADDR only."""
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Tunnel Clinic",
            address="2 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        client = Client()
        response = client.post(
            "/whatsapp/webhook/",
            data=json.dumps(self._meta_inbound_payload()),
            content_type="application/json",
            REMOTE_ADDR="127.0.0.1",
            HTTP_X_FORWARDED_FOR="157.240.0.1",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["synced"], 1)
        self.assertTrue(
            ChatMessage.objects.filter(lead=lead, is_outbound=False).exists()
        )


class YCloudWebhookTests(TestCase):
    def _ycloud_inbound_payload(self, **overrides):
        inbound = {
            "id": "inb_123",
            "wamid": "wamid.YCLOUD_INBOUND",
            "from": "+60123456789",
            "to": "+60126336529",
            "type": "text",
            "text": {"body": "Hello from YCloud"},
            "sendTime": "2026-06-14T10:00:00.000Z",
        }
        inbound.update(overrides)
        return {
            "id": "evt_inbound_1",
            "type": "whatsapp.inbound_message.received",
            "apiVersion": "v2",
            "createTime": "2026-06-14T10:00:01.000Z",
            "whatsappInboundMessage": inbound,
        }

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336529")
    def test_parse_ycloud_inbound_message(self):
        from leads.whatsapp_webhook import parse_ycloud_webhook

        parsed, failures = parse_ycloud_webhook(self._ycloud_inbound_payload())
        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(failures), 0)
        self.assertEqual(parsed[0].remote_phone, "+60123456789")
        self.assertEqual(parsed[0].text_body, "Hello from YCloud")
        self.assertFalse(parsed[0].from_me)

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_parse_ycloud_mobile_echo(self):
        from leads.whatsapp_webhook import parse_ycloud_webhook

        payload = {
            "id": "evt_out_1",
            "type": "whatsapp.message.updated",
            "apiVersion": "v2",
            "createTime": "2026-06-14T10:01:00.000Z",
            "whatsappMessage": {
                "id": "msg_1",
                "wamid": "wamid.YCLOUD_ECHO",
                "from": "+60126336429",
                "to": "+60123456789",
                "type": "text",
                "status": "sent",
                "text": {"body": "Reply from mobile app"},
                "sendTime": "2026-06-14T10:01:00.000Z",
            },
        }
        parsed, failures = parse_ycloud_webhook(payload)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(failures), 0)
        self.assertTrue(parsed[0].from_me)
        self.assertEqual(parsed[0].text_body, "Reply from mobile app")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_parse_ycloud_smb_message_echoes(self):
        from leads.whatsapp_webhook import parse_ycloud_webhook

        payload = {
            "id": "evt_smb_echo_1",
            "type": "whatsapp.smb.message.echoes",
            "apiVersion": "v2",
            "createTime": "2026-06-14T10:02:00.000Z",
            "whatsappMessage": {
                "id": "msg_smb_1",
                "wamid": "wamid.YCLOUD_SMB_ECHO",
                "from": "+60126336429",
                "to": "+60123456789",
                "type": "text",
                "status": "sent",
                "text": {"body": "Reply from Coex phone app"},
                "sendTime": "2026-06-14T10:02:00.000Z",
            },
        }
        parsed, failures = parse_ycloud_webhook(payload)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(failures), 0)
        self.assertTrue(parsed[0].from_me)
        self.assertEqual(parsed[0].remote_phone, "+60123456789")
        self.assertEqual(parsed[0].text_body, "Reply from Coex phone app")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_ycloud_template_webhook_upserts_outbound_chat(self):
        from django.utils import timezone

        from leads.chat_messages import upsert_outbound_chat_message
        from leads.whatsapp_webhook import parse_ycloud_webhook

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Template Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        WhatsAppConfig = __import__("leads.models", fromlist=["WhatsAppConfig"]).WhatsAppConfig
        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {
                "name": "say_hi",
                "status": "APPROVED",
                "language": "en_US",
                "body": "Hi- are you open today?",
            },
        ]
        config.save(update_fields=["meta_message_templates"])

        send_time = timezone.now() - timezone.timedelta(minutes=2)
        upsert_outbound_chat_message(
            lead,
            template_name="say_hi",
            body="wrong draft copy",
            meta_message_id="wamid.TEMPLATE123",
            created_at=send_time,
        )

        payload = {
            "id": "evt_tpl_1",
            "type": "whatsapp.message.updated",
            "apiVersion": "v2",
            "createTime": "2026-06-30T11:31:00.000Z",
            "whatsappMessage": {
                "id": "msg_tpl_1",
                "wamid": "wamid.TEMPLATE123",
                "from": "+60126336429",
                "to": "+60123456789",
                "type": "template",
                "status": "sent",
                "template": {"name": "say_hi", "language": {"code": "en_US"}},
                "text": {"body": "Hi- are you open today? May I know your business hours?"},
                "sendTime": "2026-06-30T11:31:00.000Z",
            },
        }
        parsed, failures = parse_ycloud_webhook(payload)
        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(failures), 0)
        self.assertEqual(parsed[0].template_name, "say_hi")
        self.assertIn("business hours", parsed[0].text_body)

        client = Client()
        response = client.post(
            "/whatsapp/webhook/",
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        chat = ChatMessage.objects.get(lead=lead, is_outbound=True)
        self.assertEqual(chat.template_name, "say_hi")
        self.assertIn("business hours", chat.body)
        self.assertNotEqual(chat.body, "wrong draft copy")
        self.assertEqual(chat.meta_message_id, "wamid.TEMPLATE123")

    def test_lead_already_received_template(self):
        from leads.chat_messages import lead_already_received_template

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Dup Template Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
        )
        self.assertFalse(lead_already_received_template(lead, "say_hi"))
        ChatMessage.objects.create(
            lead=lead,
            body="Hi~",
            is_outbound=True,
            template_name="say_hi",
        )
        self.assertTrue(lead_already_received_template(lead, "say_hi"))
        self.assertFalse(lead_already_received_template(lead, "say_hi_en"))

    def test_mark_sent_sinks_display_order_on_first_send(self):
        from leads.whatsapp_service import mark_sent

        groups = ensure_pipeline_system_groups()
        group = LeadGroup.objects.create(name="Sink Folder", sort_order=60)
        top = Lead.objects.create(
            name="Top Lead",
            address="1 Main St",
            phone_number="+60111111111",
            group=group,
            display_order=1,
        )
        lead = Lead.objects.create(
            name="Sink Lead",
            address="2 Main St",
            phone_number="+60222222222",
            group=group,
            display_order=2,
        )
        mark_sent(lead, "+60126336429", template_name="say_hi")
        lead.refresh_from_db()
        top.refresh_from_db()
        self.assertGreater(lead.display_order, top.display_order)

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_parse_ycloud_delivery_failure(self):
        from leads.whatsapp_webhook import DELIVERY_FAILED_MARKER, parse_ycloud_webhook

        payload = {
            "id": "evt_fail_1",
            "type": "whatsapp.message.updated",
            "whatsappMessage": {
                "id": "msg_fail_1",
                "wamid": "wamid.FAILED123",
                "from": "+60126336429",
                "to": "+60198765030",
                "type": "template",
                "status": "failed",
                "errorCode": "131026",
                "errorMessage": "Message undeliverable",
                "template": {"name": "say_hi", "language": {"code": "en_US"}},
            },
        }
        parsed, failures = parse_ycloud_webhook(payload)
        self.assertEqual(parsed, [])
        self.assertEqual(len(failures), 1)
        self.assertEqual(failures[0].error_message, "Message undeliverable")

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Fail Clinic",
            address="1 Main St",
            phone_number="+60198765030",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        client = Client()
        response = client.post(
            "/whatsapp/webhook/",
            data=json.dumps(payload),
            content_type="application/json",
            REMOTE_ADDR="127.0.0.1",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["delivery_failures"], 1)
        lead.refresh_from_db()
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.FAILED)
        self.assertTrue(
            LeadConversationLog.objects.filter(
                lead=lead,
                remarks__contains=DELIVERY_FAILED_MARKER,
            ).exists()
        )

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_upsert_merges_api_send_and_webhook_ids_for_same_template(self):
        from django.utils import timezone

        from leads.chat_messages import chat_messages_for_lead, upsert_outbound_chat_message
        from leads.models import WhatsAppConfig

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Merge Template Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        preview = "Hi~ are you open today? may I know your business hours?"
        config = WhatsAppConfig.load()
        config.outbound_template_name = "say_hi"
        config.meta_message_templates = [
            {
                "name": "say_hi",
                "status": "APPROVED",
                "language": "en_US",
                "body": preview,
            },
        ]
        config.save(update_fields=["outbound_template_name", "meta_message_templates"])
        send_time = timezone.now() - timezone.timedelta(minutes=7)
        upsert_outbound_chat_message(
            lead,
            template_name="say_hi",
            body=preview,
            meta_message_id="ycloud_msg_1",
            created_at=send_time,
        )
        upsert_outbound_chat_message(
            lead,
            template_name="say_hi",
            body=preview,
            meta_message_id="wamid.TEMPLATE123",
            created_at=timezone.now(),
        )

        messages = chat_messages_for_lead(lead)
        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0].template_name, "say_hi")
        self.assertEqual(messages[0].meta_message_id, "wamid.TEMPLATE123")
        self.assertEqual(messages[0].body, preview)

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_upsert_merges_free_text_api_and_webhook_ids(self):
        from django.utils import timezone

        from leads.chat_messages import chat_messages_for_lead, upsert_outbound_chat_message

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Merge Free Text Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        body = "We open at 9am tomorrow."
        send_time = timezone.now() - timezone.timedelta(minutes=3)
        upsert_outbound_chat_message(
            lead,
            body=body,
            meta_message_id="ycloud_text_1",
            template_name="",
            created_at=send_time,
        )
        upsert_outbound_chat_message(
            lead,
            body=body,
            meta_message_id="wamid.FREE_TEXT_ECHO",
            template_name="",
            created_at=timezone.now(),
        )

        messages = chat_messages_for_lead(lead)
        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0].body, body)
        self.assertEqual(messages[0].template_name, "")
        self.assertEqual(messages[0].meta_message_id, "wamid.FREE_TEXT_ECHO")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_ycloud_smb_echo_webhook_post_syncs_outbound(self):
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Coex Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        payload = {
            "id": "evt_smb_echo_2",
            "type": "whatsapp.smb.message.echoes",
            "apiVersion": "v2",
            "createTime": "2026-06-14T10:02:00.000Z",
            "whatsappMessage": {
                "id": "msg_smb_2",
                "wamid": "wamid.YCLOUD_SMB_ECHO_POST",
                "from": "+60126336429",
                "to": "+60123456789",
                "type": "text",
                "text": {"body": "Coex phone reply logged"},
                "sendTime": "2026-06-14T10:02:00.000Z",
            },
        }
        client = Client()
        response = client.post(
            "/whatsapp/webhook/",
            data=json.dumps(payload),
            content_type="application/json",
            REMOTE_ADDR="127.0.0.1",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["synced"], 1)
        chat = ChatMessage.objects.get(lead=lead, is_outbound=True)
        self.assertEqual(chat.body, "Coex phone reply logged")
        self.assertEqual(chat.template_name, "")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_business_app_free_text_not_labeled_as_template(self):
        from leads.chat_messages import chat_messages_for_lead, upsert_outbound_chat_message

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Free Text Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        upsert_outbound_chat_message(
            lead,
            body="Thanks, we open at 9am tomorrow.",
            meta_message_id="wamid.FREE_TEXT_ECHO",
            template_name="",
        )
        messages = chat_messages_for_lead(lead)
        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0].body, "Thanks, we open at 9am tomorrow.")
        self.assertEqual(messages[0].template_name, "")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_sync_logs_do_not_retag_free_text_as_template(self):
        from django.utils import timezone

        from leads.chat_messages import chat_messages_for_lead, record_outbound_chat_message
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import OFFICIAL_API_MARKER

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Retag Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        config = WhatsAppConfig.load()
        config.outbound_template_name = "say_hi"
        config.meta_message_templates = [
            {
                "name": "say_hi",
                "status": "APPROVED",
                "language": "en_US",
                "body": "Hi~ are you open today? may I know your business hours?",
            },
        ]
        config.save(update_fields=["outbound_template_name", "meta_message_templates"])

        record_outbound_chat_message(
            lead,
            template_name="say_hi",
            body="Hi~ are you open today? may I know your business hours?",
        )
        record_outbound_chat_message(
            lead,
            template_name="",
            body="We open at 9am — reply from Business app",
        )
        LeadConversationLog.objects.create(
            lead=lead,
            conversation_date=timezone.now().date(),
            remarks=f"{OFFICIAL_API_MARKER} Template queued by YCloud for 6012XXXX789",
        )

        messages = chat_messages_for_lead(lead)
        free_text = [m for m in messages if "9am" in m.body]
        self.assertEqual(len(free_text), 1)
        self.assertEqual(free_text[0].template_name, "")

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336429")
    def test_chat_messages_for_lead_imports_agent_log_and_relabels_you(self):
        from django.utils import timezone

        from leads.chat_messages import (
            chat_messages_for_lead,
            outbound_message_is_template,
        )
        from leads.models import WhatsAppConfig

        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="Sync Button Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        config = WhatsAppConfig.load()
        config.outbound_template_name = "say_hi"
        config.meta_message_templates = [
            {
                "name": "say_hi",
                "status": "APPROVED",
                "language": "en_US",
                "body": "Hi~ are you open today? may I know your business hours?",
            },
        ]
        config.save(update_fields=["outbound_template_name", "meta_message_templates"])

        LeadConversationLog.objects.create(
            lead=lead,
            conversation_date=timezone.now().date(),
            remarks="[WhatsApp · agent] We open at 9am — reply from Business app",
        )

        messages = chat_messages_for_lead(lead)
        self.assertEqual(len(messages), 1)

        msg = messages[0]
        self.assertEqual(msg.body, "We open at 9am — reply from Business app")
        self.assertEqual(msg.template_name, "")
        self.assertFalse(outbound_message_is_template(msg))

    @override_settings(WHATSAPP_FROM_NUMBER="+60126336529")
    def test_ycloud_webhook_post_syncs_inbound(self):
        groups = ensure_pipeline_system_groups()
        lead = Lead.objects.create(
            name="YCloud Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        client = Client()
        response = client.post(
            "/whatsapp/webhook/",
            data=json.dumps(self._ycloud_inbound_payload()),
            content_type="application/json",
            REMOTE_ADDR="127.0.0.1",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["synced"], 1)
        chat = ChatMessage.objects.get(lead=lead, is_outbound=False)
        self.assertEqual(chat.body, "Hello from YCloud")


class PhoneDeduplicationTests(TestCase):
    def test_phone_exists_matches_primary_and_json_numbers(self):
        Lead.objects.create(
            name="Existing",
            address="9 Road",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
        )
        self.assertTrue(phone_exists_in_database("0123456789"))
        self.assertTrue(phone_exists_in_database("+60123456789"))
        self.assertFalse(phone_exists_in_database("+60999998888"))


class ChatFreeTextSendTests(TestCase):
    def setUp(self):
        groups = ensure_pipeline_system_groups()
        self.lead = Lead.objects.create(
            name="Chat Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=groups["uncategorized"],
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )

    @patch("leads.views.send_free_text_to_lead")
    def test_send_free_text_returns_outbound_bubble(self, mock_send):
        from leads.models import ChatMessage

        msg = ChatMessage.objects.create(
            lead=self.lead,
            body="Thanks for your reply!",
            is_outbound=True,
            template_name="",
        )
        mock_send.return_value = (True, "", msg)

        client = Client()
        response = client.post(
            reverse("send_free_text", kwargs={"pk": self.lead.pk}),
            {"message": "Thanks for your reply!"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Thanks for your reply!", response.content.decode())
        mock_send.assert_called_once_with(self.lead, "Thanks for your reply!")

    def test_send_free_text_rejects_empty_body(self):
        client = Client()
        response = client.post(
            reverse("send_free_text", kwargs={"pk": self.lead.pk}),
            {"message": "   "},
        )
        self.assertEqual(response.status_code, 400)


class WhatsAppBatchScheduleTests(TestCase):
    def _make_pending_lead(self, name: str, phone: str = "+60123456789") -> Lead:
        return Lead.objects.create(
            name=name,
            address=f"{name} St",
            group=get_or_create_uncategorized_group(),
            phone_number=phone,
            whatsapp_status=Lead.WhatsappStatus.PENDING,
        )

    def test_dispatch_pending_batch_sends_oldest_first_up_to_limit(self):
        from leads.whatsapp_service import dispatch_pending_batch

        first = self._make_pending_lead("Alpha")
        second = self._make_pending_lead("Beta")
        self._make_pending_lead("Gamma")

        sent_order = []

        def fake_send(lead, *, priority=False, template_name=None):
            sent_order.append(lead.pk)
            lead.whatsapp_status = Lead.WhatsappStatus.SENT
            lead.save(update_fields=["whatsapp_status"])
            return True, ""

        with patch("leads.whatsapp_service.send_text_to_lead", side_effect=fake_send):
            sent = dispatch_pending_batch(2)

        self.assertEqual(sent, 2)
        self.assertEqual(sent_order, [first.pk, second.pk])
        self.assertEqual(
            Lead.objects.filter(whatsapp_status=Lead.WhatsappStatus.PENDING).count(),
            1,
        )

    def test_run_due_scheduled_batches_executes_only_due(self):
        from django.utils import timezone
        from datetime import timedelta

        from leads.models import WhatsAppBatchSchedule
        from leads.whatsapp_service import run_due_scheduled_batches

        due = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() - timedelta(minutes=1),
        )
        future = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=1),
        )
        # Two leads assigned to the due batch, one assigned to the future batch.
        for i in range(2):
            lead = self._make_pending_lead(f"Due{i}")
            lead.whatsapp_batches.add(due)
        later = self._make_pending_lead("Later")
        later.whatsapp_batches.add(future)

        def fake_send(lead, *, priority=False, template_name=None):
            lead.whatsapp_status = Lead.WhatsappStatus.SENT
            lead.save(update_fields=["whatsapp_status"])
            return True, ""

        with patch("leads.whatsapp_service.send_text_to_lead", side_effect=fake_send):
            summary = run_due_scheduled_batches()

        due.refresh_from_db()
        future.refresh_from_db()
        self.assertEqual(summary["batches_run"], 1)
        self.assertEqual(summary["leads_sent"], 2)
        self.assertEqual(due.status, WhatsAppBatchSchedule.Status.COMPLETED)
        self.assertEqual(due.sent_count, 2)
        self.assertEqual(future.status, WhatsAppBatchSchedule.Status.PENDING)
        self.assertEqual(
            Lead.objects.filter(whatsapp_status=Lead.WhatsappStatus.PENDING).count(),
            1,
        )

    def test_schedule_batch_view_creates_future_batch(self):
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        target = (timezone.localtime() + timedelta(days=1)).replace(
            second=0, microsecond=0
        )
        client = Client()
        response = client.post(
            reverse("whatsapp_schedule_batch"),
            {
                "scheduled_date": target.strftime("%Y-%m-%d"),
                "scheduled_time": target.strftime("%H:%M"),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(WhatsAppBatchSchedule.objects.count(), 1)
        batch = WhatsAppBatchSchedule.objects.first()
        self.assertEqual(batch.status, WhatsAppBatchSchedule.Status.PENDING)

    def test_schedule_batch_view_rejects_past_datetime(self):
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        target = timezone.localtime() - timedelta(days=1)
        client = Client()
        response = client.post(
            reverse("whatsapp_schedule_batch"),
            {
                "scheduled_date": target.strftime("%Y-%m-%d"),
                "scheduled_time": target.strftime("%H:%M"),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(WhatsAppBatchSchedule.objects.count(), 0)

    def test_cancel_batch_view_cancels_pending(self):
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        batch = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=2),
        )
        client = Client()
        response = client.post(reverse("whatsapp_cancel_batch", kwargs={"pk": batch.pk}))
        self.assertEqual(response.status_code, 200)
        batch.refresh_from_db()
        self.assertEqual(batch.status, WhatsAppBatchSchedule.Status.CANCELLED)

    def test_bulk_assign_batch_assigns_to_existing_batch(self):
        import json as _json
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        batch = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=3),
        )
        a = self._make_pending_lead("Aa")
        b = self._make_pending_lead("Bb")
        client = Client()
        response = client.post(
            reverse("leads_bulk_assign_batch"),
            data=_json.dumps({"ids": [a.pk, b.pk], "batch_id": batch.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["updated"], 2)
        a.refresh_from_db()
        b.refresh_from_db()
        self.assertTrue(a.whatsapp_batches.filter(pk=batch.pk).exists())
        self.assertTrue(b.whatsapp_batches.filter(pk=batch.pk).exists())

    def test_bulk_assign_batch_creates_new_batch(self):
        import json as _json
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        target = (timezone.localtime() + timedelta(days=1)).replace(
            second=0, microsecond=0
        )
        a = self._make_pending_lead("Cc")
        client = Client()
        response = client.post(
            reverse("leads_bulk_assign_batch"),
            data=_json.dumps(
                {
                    "ids": [a.pk],
                    "batch_id": "new",
                    "new_batch": {
                        "scheduled_date": target.strftime("%Y-%m-%d"),
                        "scheduled_time": target.strftime("%H:%M"),
                        "outbound_template_name": "just_to_say_hi",
                    },
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(WhatsAppBatchSchedule.objects.count(), 1)
        a.refresh_from_db()
        self.assertTrue(a.whatsapp_batches.exists())

    def test_dequeue_clears_pending_batch_assignment(self):
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        batch = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=2),
        )
        lead = self._make_pending_lead("Dequeued")
        lead.whatsapp_batches.add(batch)

        client = Client()
        response = client.post(reverse("dequeue_lead", kwargs={"pk": lead.pk}))
        self.assertEqual(response.status_code, 200)

        lead.refresh_from_db()
        self.assertEqual(lead.whatsapp_status, Lead.WhatsappStatus.IDLE)
        self.assertFalse(lead.whatsapp_batches.filter(pk=batch.pk).exists())

    def test_dequeue_keeps_completed_batch_history(self):
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        done = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() - timedelta(hours=2),
            status=WhatsAppBatchSchedule.Status.COMPLETED,
        )
        lead = self._make_pending_lead("KeepHistory")
        lead.whatsapp_batches.add(done)

        client = Client()
        response = client.post(reverse("dequeue_lead", kwargs={"pk": lead.pk}))
        self.assertEqual(response.status_code, 200)

        lead.refresh_from_db()
        self.assertTrue(lead.whatsapp_batches.filter(pk=done.pk).exists())

    def test_bulk_dequeue_removes_pending_leads_from_queue(self):
        import json as _json
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        batch = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=2),
        )
        pending_a = self._make_pending_lead("BulkA")
        pending_b = self._make_pending_lead("BulkB")
        processing = self._make_pending_lead("Processing")
        processing.whatsapp_status = Lead.WhatsappStatus.PROCESSING
        processing.save(update_fields=["whatsapp_status"])
        pending_a.whatsapp_batches.add(batch)
        pending_b.whatsapp_batches.add(batch)

        client = Client()
        response = client.post(
            reverse("leads_bulk_dequeue"),
            data=_json.dumps({"ids": [pending_a.pk, pending_b.pk, processing.pk]}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["updated"], 2)
        self.assertEqual(data["skipped"], 1)

        pending_a.refresh_from_db()
        pending_b.refresh_from_db()
        processing.refresh_from_db()
        self.assertEqual(pending_a.whatsapp_status, Lead.WhatsappStatus.IDLE)
        self.assertEqual(pending_b.whatsapp_status, Lead.WhatsappStatus.IDLE)
        self.assertEqual(processing.whatsapp_status, Lead.WhatsappStatus.PROCESSING)
        self.assertFalse(pending_a.whatsapp_batches.filter(pk=batch.pk).exists())
        self.assertFalse(pending_b.whatsapp_batches.filter(pk=batch.pk).exists())

    def test_bulk_assign_batch_skips_leads_already_in_pending_batch(self):
        import json as _json
        from datetime import timedelta

        from django.utils import timezone

        from leads.models import WhatsAppBatchSchedule

        existing = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=1),
        )
        target = WhatsAppBatchSchedule.objects.create(
            scheduled_at=timezone.now() + timedelta(hours=2),
        )
        already = self._make_pending_lead("Already")
        already.whatsapp_batches.add(existing)
        fresh = self._make_pending_lead("Fresh")

        client = Client()
        response = client.post(
            reverse("leads_bulk_assign_batch"),
            data=_json.dumps({"ids": [already.pk, fresh.pk], "batch_id": target.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["updated"], 1)
        self.assertEqual(payload["skipped"], 1)
        # The already-queued lead stays out of the target batch.
        self.assertFalse(already.whatsapp_batches.filter(pk=target.pk).exists())
        self.assertTrue(fresh.whatsapp_batches.filter(pk=target.pk).exists())


class WhatsAppMetaTemplateSyncTests(TestCase):
    def test_meta_template_choices_uses_synced_catalog(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import meta_template_choices_for_ui

        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {"name": "custom_greeting", "status": "APPROVED", "language": "en", "body": "Hi"},
            {"name": "just_to_say_hi", "status": "APPROVED", "language": "en", "body": "Hello"},
        ]
        config.save(update_fields=["meta_message_templates"])

        choices = dict(meta_template_choices_for_ui())
        self.assertIn("custom_greeting", choices)
        self.assertIn("just_to_say_hi", choices)
        self.assertIn("(Default", choices["just_to_say_hi"])

    @patch("leads.whatsapp_service.fetch_meta_message_templates_from_api")
    def test_sync_meta_message_templates_persists_catalog(self, mock_fetch):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import sync_meta_message_templates_to_config

        mock_fetch.return_value = [
            {"name": "just_to_say_hi", "status": "APPROVED", "language": "en", "body": "Hello"},
            {"name": "promo_v2", "status": "APPROVED", "language": "en", "body": "Promo"},
        ]
        count, error = sync_meta_message_templates_to_config()
        self.assertIsNone(error)
        self.assertEqual(count, 2)

        config = WhatsAppConfig.load()
        self.assertEqual(len(config.meta_message_templates), 2)
        self.assertIsNotNone(config.meta_templates_synced_at)

    def test_meta_template_language_preserves_en_us(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import build_meta_template_payload, meta_template_language_for_name

        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {
                "name": "say_hi",
                "status": "APPROVED",
                "language": "en_US",
                "body": "Hi- are you open today?",
            },
        ]
        config.outbound_template_name = "say_hi"
        config.save(update_fields=["meta_message_templates", "outbound_template_name"])

        self.assertEqual(meta_template_language_for_name("say_hi"), "en_US")
        payload = build_meta_template_payload(
            Lead.objects.create(name="Test Clinic", phone_number="+60123456789"),
            template_name="say_hi",
        )
        self.assertEqual(payload["template"]["language"]["code"], "en_US")

    def test_get_force_send_template_name_falls_back_to_outbound(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import get_force_send_template_name

        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {"name": "say_hi", "status": "APPROVED", "language": "en_US", "body": "Hi"},
            {"name": "say_hi_en", "status": "APPROVED", "language": "en", "body": "Hello"},
        ]
        config.outbound_template_name = "say_hi"
        config.force_send_template_name = ""
        config.save(
            update_fields=[
                "meta_message_templates",
                "outbound_template_name",
                "force_send_template_name",
            ]
        )
        self.assertEqual(get_force_send_template_name(), "say_hi")

        config.force_send_template_name = "say_hi_en"
        config.save(update_fields=["force_send_template_name"])
        self.assertEqual(get_force_send_template_name(), "say_hi_en")

    def test_known_meta_template_names_requires_synced_catalog(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import known_meta_template_names

        config = WhatsAppConfig.load()
        config.meta_message_templates = []
        config.save(update_fields=["meta_message_templates"])
        self.assertEqual(known_meta_template_names(), frozenset())

    def test_validate_outbound_template_name_accepts_synced_name(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import validate_outbound_template_name

        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {"name": "hello_clinic", "status": "APPROVED", "language": "en", "body": "Hi"},
        ]
        config.save(update_fields=["meta_message_templates"])

        ok, err = validate_outbound_template_name("hello_clinic")
        self.assertTrue(ok)
        self.assertEqual(err, "")

    def test_validate_outbound_template_name_rejects_unknown_name(self):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import validate_outbound_template_name

        config = WhatsAppConfig.load()
        config.meta_message_templates = [
            {"name": "hello_clinic", "status": "APPROVED", "language": "en", "body": "Hi"},
        ]
        config.save(update_fields=["meta_message_templates"])

        ok, err = validate_outbound_template_name("just_to_say_hi")
        self.assertFalse(ok)
        self.assertIn("not approved on YCloud", err)

    @patch("leads.whatsapp_service.fetch_meta_message_templates_from_api")
    def test_normalize_outbound_template_name_uses_catalog_default(self, mock_fetch):
        from leads.models import WhatsAppConfig
        from leads.whatsapp_service import normalize_outbound_template_name, sync_meta_message_templates_to_config

        mock_fetch.return_value = [
            {"name": "hello_clinic", "status": "APPROVED", "language": "en", "body": "Hi", "wabaId": "123"},
        ]
        sync_meta_message_templates_to_config()
        config = WhatsAppConfig.load()
        config.outbound_template_name = "just_to_say_hi"
        config.save(update_fields=["outbound_template_name"])

        self.assertEqual(normalize_outbound_template_name("just_to_say_hi"), "hello_clinic")

    @patch("leads.views.sync_meta_message_templates_to_config")
    def test_refresh_meta_templates_view_returns_toast_and_oob_field(self, mock_sync):
        mock_sync.return_value = (3, None)

        client = Client()
        response = client.post(reverse("whatsapp_refresh_meta_templates"))
        self.assertEqual(response.status_code, 200)
        body = response.content.decode()
        self.assertIn("Synced 3 approved template(s) from YCloud", body)
        self.assertIn('id="outbound-template-field"', body)
        self.assertIn("hx-swap-oob", body)
        mock_sync.assert_called_once()

    @patch("leads.views.sync_meta_message_templates_to_config")
    def test_refresh_meta_templates_view_shows_error_toast(self, mock_sync):
        mock_sync.return_value = (0, "Token expired")

        client = Client()
        response = client.post(reverse("whatsapp_refresh_meta_templates"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("Template sync failed: Token expired", response.content.decode())


class YCloudWabaResolveTests(TestCase):
    @patch("leads.ycloud_service.httpx.Client")
    def test_resolve_sending_waba_id_from_phone_numbers_api(self, mock_client_cls):
        from leads.ycloud_service import resolve_sending_waba_id

        mock_response = mock_client_cls.return_value.__enter__.return_value.get
        mock_response.return_value.status_code = 200
        mock_response.return_value.json.return_value = {
            "items": [
                {
                    "phoneNumber": "+60126336429",
                    "wabaId": "1478974178167699",
                }
            ],
            "page": {"length": 1, "limit": 100},
        }

        with self.settings(WHATSAPP_FROM_NUMBER="+60126336429", YCLOUD_WABA_ID="1470974178167699"):
            waba = resolve_sending_waba_id(refresh=True)
        self.assertEqual(waba, "1478974178167699")


class ClinicUpdatePhoneTests(TestCase):
    def setUp(self):
        groups = ensure_pipeline_system_groups()
        self.group = LeadGroup.objects.create(name="Test Folder", sort_order=50)
        self.lead = Lead.objects.create(
            name="Phone Change Clinic",
            address="1 Main St",
            phone_number="+60123456789",
            phone_numbers=["+60123456789"],
            group=self.group,
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        ChatMessage.objects.create(
            lead=self.lead,
            body="Hi",
            is_outbound=True,
            template_name="say_hi",
        )

    def test_clinic_update_phone_resets_whatsapp_dispatch_state(self):
        client = Client()
        response = client.patch(
            reverse("clinic_update", kwargs={"pk": self.lead.pk}),
            data=json.dumps(
                {
                    "name": self.lead.name,
                    "phone_numbers": ["+60198765432"],
                    "address": self.lead.address,
                    "category": "unknown",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["phone_numbers"], ["+60198765432"])
        self.assertEqual(data["whatsapp_status"], Lead.WhatsappStatus.IDLE)
        self.assertFalse(data["whatsapp_dispatched"])
        self.assertIn("lead-force-send-btn", data["grid_bottom_actions_html"])

        self.lead.refresh_from_db()
        self.assertEqual(self.lead.phone_number, "+60198765432")
        self.assertEqual(self.lead.whatsapp_status, Lead.WhatsappStatus.IDLE)
        self.assertIsNone(self.lead.whatsapp_sent_at)
        self.assertFalse(ChatMessage.objects.filter(lead=self.lead).exists())

    def test_primary_phone_uses_phone_numbers_over_stale_phone_number(self):
        from leads.whatsapp_service import build_meta_template_payload, primary_phone

        self.lead.phone_number = "+60111111111"
        self.lead.phone_numbers = ["+60222222222"]
        self.assertEqual(primary_phone(self.lead), "+60222222222")

        payload = build_meta_template_payload(self.lead, template_name="say_hi")
        self.assertEqual(payload["to"], "+60222222222")

    def test_force_send_button_shown_when_status_sent(self):
        from django.test import RequestFactory

        from leads.views import _force_send_grid_response

        self.lead.whatsapp_status = Lead.WhatsappStatus.SENT
        request = RequestFactory().post(
            "/",
            data={"group_id": str(self.group.pk)},
        )
        response = _force_send_grid_response(
            request,
            self.lead,
            ok=True,
        )
        self.assertIn("lead-force-send-btn", response.content.decode())

    @patch("leads.views.send_text_to_lead")
    def test_force_send_duplicate_prompts_resend(self, mock_send):
        from leads.models import WhatsAppConfig

        config = WhatsAppConfig.load()
        config.force_send_template_name = "say_hi"
        config.save(update_fields=["force_send_template_name"])
        ChatMessage.objects.create(
            lead=self.lead,
            body="Hi~ are you open today?",
            is_outbound=True,
            template_name="say_hi",
        )
        client = Client()
        with self.settings(WHATSAPP_FROM_NUMBER="+60126336429", YCLOUD_API_KEY="test"):
            response = client.post(
                reverse("whatsapp_force_send", kwargs={"pk": self.lead.pk}),
                data={"group_id": str(self.group.pk)},
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_not_called()
        trigger = json.loads(response["HX-Trigger"])
        self.assertIn("forceSendDuplicatePrompt", trigger)
        self.assertEqual(trigger["forceSendDuplicatePrompt"]["leadId"], self.lead.pk)
        self.assertNotIn("leadCardSink", trigger)

    @patch("leads.views.send_text_to_lead")
    def test_force_send_confirm_duplicate_resends(self, mock_send):
        from leads.models import WhatsAppConfig

        mock_send.return_value = (True, "accepted")
        config = WhatsAppConfig.load()
        config.force_send_template_name = "say_hi"
        config.save(update_fields=["force_send_template_name"])
        ChatMessage.objects.create(
            lead=self.lead,
            body="Hi~ are you open today?",
            is_outbound=True,
            template_name="say_hi",
        )
        client = Client()
        with self.settings(WHATSAPP_FROM_NUMBER="+60126336429", YCLOUD_API_KEY="test"):
            response = client.post(
                reverse("whatsapp_force_send", kwargs={"pk": self.lead.pk}),
                data={
                    "group_id": str(self.group.pk),
                    "confirm_duplicate": "1",
                },
            )
        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once()

    def test_force_send_success_triggers_dispatched_and_sink(self):
        from django.test import RequestFactory
        from django.utils import timezone

        from leads.views import _force_send_grid_response

        self.lead.whatsapp_sent_at = timezone.now()
        self.lead.whatsapp_status = Lead.WhatsappStatus.SENT
        self.lead.save(update_fields=["whatsapp_sent_at", "whatsapp_status"])
        request = RequestFactory().post(
            "/",
            data={"group_id": str(self.group.pk)},
        )
        response = _force_send_grid_response(
            request,
            self.lead,
            ok=True,
            sink_card=True,
        )
        trigger = json.loads(response["HX-Trigger"])
        self.assertEqual(trigger["leadCardDispatched"], self.lead.pk)
        self.assertEqual(trigger["leadCardSink"], self.lead.pk)
        body = response.content.decode()
        self.assertIn("hx-swap-oob", body)
        self.assertIn("clinic-card--dispatched", body)


class DailyReportTests(TestCase):
    def setUp(self):
        from datetime import datetime, time

        from django.utils import timezone

        from leads.whatsapp_service import campaign_timezone

        self.tz = campaign_timezone()
        self.today = timezone.now().astimezone(self.tz).date()
        start = timezone.make_aware(datetime.combine(self.today, time.min), self.tz)
        self.lead_sent = Lead.objects.create(
            name="Alpha Clinic",
            address="1 Main St",
            group=get_or_create_uncategorized_group(),
            phone_number="+60111111111",
            search_state="Selangor",
            search_city="Petaling Jaya",
            whatsapp_status=Lead.WhatsappStatus.SENT,
            whatsapp_sent_at=start,
        )
        ChatMessage.objects.create(
            lead=self.lead_sent,
            body="Hi",
            is_outbound=True,
            template_name="say_hi",
            created_at=start,
        )
        self.lead_reply = Lead.objects.create(
            name="Beta Clinic",
            address="2 Main St",
            group=get_or_create_uncategorized_group(),
            phone_number="+60222222222",
            whatsapp_status=Lead.WhatsappStatus.SENT,
        )
        ChatMessage.objects.create(
            lead=self.lead_reply,
            body="Hello back",
            is_outbound=False,
            created_at=start,
        )

    def test_reports_excludes_leads_without_outbound_that_day(self):
        from django.utils import timezone

        from leads.views import _daily_report_leads

        start = timezone.make_aware(
            __import__("datetime").datetime.combine(self.today, __import__("datetime").time.min),
            self.tz,
        )
        log_only = Lead.objects.create(
            name="Log Only Clinic",
            address="9 Main St",
            group=get_or_create_uncategorized_group(),
            phone_number="+60999999999",
        )
        LeadConversationLog.objects.create(
            lead=log_only,
            conversation_date=self.today,
            remarks="Phone number updated.",
        )
        leads = _daily_report_leads(self.today)
        names = {lead.name for lead in leads}
        self.assertIn("Alpha Clinic", names)
        self.assertNotIn("Log Only Clinic", names)

    def test_reports_shows_active_when_inbound_and_failed(self):
        from django.utils import timezone

        from leads.views import _daily_report_leads

        start = timezone.make_aware(
            __import__("datetime").datetime.combine(self.today, __import__("datetime").time.min),
            self.tz,
        )
        failed_active = Lead.objects.create(
            name="Failed But Active",
            address="8 Main St",
            group=get_or_create_uncategorized_group(),
            phone_number="+60888888888",
            whatsapp_status=Lead.WhatsappStatus.FAILED,
        )
        ChatMessage.objects.create(
            lead=failed_active,
            body="Hi",
            is_outbound=True,
            template_name="say_hi",
            created_at=start,
        )
        ChatMessage.objects.create(
            lead=failed_active,
            body="Thanks",
            is_outbound=False,
            created_at=start,
        )
        leads = {lead.name: lead for lead in _daily_report_leads(self.today)}
        self.assertEqual(leads["Failed But Active"].report_status_display, "Active")

    def test_reports_page_shows_daily_dashboard(self):
        client = Client()
        response = client.get(
            reverse("reports"),
            {"date": self.today.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("Daily reports", html)
        self.assertIn("Alpha Clinic", html)
        self.assertNotIn("Beta Clinic", html)
        self.assertIn("First sends", html)
        self.assertIn("Selangor / Petaling Jaya", html)
        self.assertIn("Click for state breakdown", html)

    def test_reports_page_shows_state_breakdown_for_metric(self):
        client = Client()
        response = client.get(
            reverse("reports"),
            {"date": self.today.isoformat(), "metric": "first_sends"},
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn('id="report-breakdown-drawer"', html)
        self.assertIn('id="report-breakdown-data"', html)
        self.assertIn('"first_sends"', html)
        self.assertIn("Selangor", html)

    def test_daily_report_location_and_state_breakdown_helpers(self):
        from leads.views import (
            _daily_report_leads,
            _daily_report_location_display,
            _daily_report_state_breakdown,
        )

        leads = {lead.name: lead for lead in _daily_report_leads(self.today)}
        self.assertEqual(
            _daily_report_location_display(leads["Alpha Clinic"]),
            "Selangor / Petaling Jaya",
        )
        self.assertEqual(
            leads["Alpha Clinic"].report_location_display,
            "Selangor / Petaling Jaya",
        )
        breakdown = _daily_report_state_breakdown(self.today, "first_sends")
        self.assertEqual(breakdown, [{"state": "Selangor", "count": 1}])

    def test_daily_report_export_xlsx(self):
        client = Client()
        response = client.get(
            reverse("daily_report_export_xlsx"),
            {"date": self.today.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(i for i, row in enumerate(rows) if row[0] == "Name")
        self.assertEqual(
            rows[header_idx],
            (
                "Name",
                "State / area",
                "Contact number",
                "First send today",
                "Outbound",
                "Inbound",
                "Status",
            ),
        )
        names = {row[0] for row in rows[header_idx + 1 :]}
        self.assertIn("Alpha Clinic", names)
        self.assertNotIn("Beta Clinic", names)
        alpha_row = next(row for row in rows[header_idx + 1 :] if row[0] == "Alpha Clinic")
        self.assertEqual(alpha_row[1], "Selangor / Petaling Jaya")

        ws_states = wb["By state"]
        state_rows = list(ws_states.iter_rows(values_only=True))
        self.assertEqual(state_rows[0], ("Metric", "State", "Count"))
        self.assertIn(("First sends", "Selangor", 1), state_rows)

    def test_monthly_report_export_xlsx(self):
        client = Client()
        month = self.today.strftime("%Y-%m")
        response = client.get(
            reverse("monthly_report_export_xlsx"),
            {"month": month},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(wb.active.title, "Monthly summary")
        summary_rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(summary_rows[0][:2], ("Month", month))
        summary_pairs = [(row[0], row[1]) for row in summary_rows if row[0]]
        self.assertIn(("First messages sent", 1), summary_pairs)
        self.assertIn("By day", wb.sheetnames)
        self.assertIn("By state", wb.sheetnames)
        day_rows = list(wb["By day"].iter_rows(values_only=True))
        self.assertEqual(day_rows[0][0], "Date")
        self.assertTrue(any(row[0] == self.today.isoformat() for row in day_rows[1:]))


class CategoryRuleManagementTests(TestCase):
    def test_category_rules_page_lists_rules(self):
        CategoryRule.objects.create(
            match_phrase="dental",
            category=Lead.Category.DENTAL,
            priority=10,
        )
        client = Client()
        response = client.get(reverse("category_rules"))
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"dental", response.content)

    def test_category_types_fragment_returns_manage_html(self):
        client = Client()
        response = client.get(reverse("category_types_fragment"))
        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn("data-category-type-form", html)
        self.assertIn("Unknown", html)

    def test_category_type_save_via_fragment_header(self):
        client = Client()
        response = client.post(
            reverse("category_type_save"),
            data={
                "label": "Veterinary",
                "slug": "vet",
                "sort_order": "50",
            },
            HTTP_X_CATEGORY_TYPES_FRAGMENT="1",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Veterinary", response.content)
        self.assertTrue(LeadCategoryType.objects.filter(slug="vet").exists())

    def test_category_type_save_and_delete(self):
        client = Client()
        create = client.post(
            reverse("category_type_save"),
            data={
                "label": "Pilates",
                "slug": "pilates",
                "sort_order": "80",
            },
        )
        self.assertEqual(create.status_code, 302)
        cat_type = LeadCategoryType.objects.get(slug="pilates")
        self.assertEqual(cat_type.label, "Pilates")

        update = client.post(
            reverse("category_type_save"),
            data={
                "id": str(cat_type.pk),
                "label": "Pilates Studio",
                "slug": "pilates",
                "sort_order": "75",
            },
        )
        self.assertEqual(update.status_code, 302)
        cat_type.refresh_from_db()
        self.assertEqual(cat_type.label, "Pilates Studio")

        delete = client.post(reverse("category_type_delete", kwargs={"pk": cat_type.pk}))
        self.assertEqual(delete.status_code, 302)
        self.assertFalse(LeadCategoryType.objects.filter(pk=cat_type.pk).exists())

    def test_category_rule_save_and_delete(self):
        client = Client()
        create = client.post(
            reverse("category_rule_save"),
            data={
                "match_phrase": "gym",
                "category": Lead.Category.FITNESS,
                "priority": "50",
            },
        )
        self.assertEqual(create.status_code, 302)
        rule = CategoryRule.objects.get(match_phrase="gym")
        self.assertEqual(rule.category, Lead.Category.FITNESS)

        update = client.post(
            reverse("category_rule_save"),
            data={
                "id": str(rule.pk),
                "match_phrase": "fitness",
                "category": Lead.Category.FITNESS,
                "priority": "20",
            },
        )
        self.assertEqual(update.status_code, 302)
        rule.refresh_from_db()
        self.assertEqual(rule.match_phrase, "fitness")
        self.assertEqual(rule.priority, 20)

        delete = client.post(reverse("category_rule_delete", kwargs={"pk": rule.pk}))
        self.assertEqual(delete.status_code, 302)
        self.assertFalse(CategoryRule.objects.filter(pk=rule.pk).exists())


class SerperHuntPaginationTests(TestCase):
    @override_settings(SERPER_API_KEY="test-key", HUNT_MAX_LIMIT=100)
    @patch("leads.services.requests.post")
    def test_fetch_paginates_when_limit_above_page_size(self, mock_post):
        from leads.services import fetch_leads_from_serper

        page1 = [
            {
                "title": f"Biz {i}",
                "address": f"St {i}",
                "phoneNumber": f"+6012{i:07d}",
                "latitude": 1.49 + i * 0.001,
                "longitude": 103.74 + i * 0.001,
            }
            for i in range(20)
        ]
        page2 = [
            {
                "title": f"Biz {i}",
                "address": f"St {i}",
                "phoneNumber": f"+6013{i:07d}",
                "latitude": 1.50 + i * 0.001,
                "longitude": 103.75 + i * 0.001,
            }
            for i in range(20, 35)
        ]

        def _resp(places):
            resp = Mock()
            resp.raise_for_status = Mock()
            resp.json.return_value = {
                "places": places,
                "searchParameters": {"ll": "@1.4927,103.7414,13z"},
            }
            return resp

        mock_post.side_effect = [_resp(page1), _resp(page2)]

        result = fetch_leads_from_serper(
            "Kuala Lumpur",
            "",
            num=40,
            shop_keyword="dental clinic",
            state="Selangor",
            country="Malaysia",
        )

        self.assertEqual(mock_post.call_count, 2)
        second_payload = mock_post.call_args_list[1][1]["json"]
        self.assertEqual(second_payload["page"], 2)
        self.assertEqual(second_payload["ll"], "@1.4927,103.7414,13z")
        self.assertEqual(result.places_seen, 35)
        self.assertEqual(result.created, 35)
        self.assertEqual(Lead.objects.count(), 35)

    @override_settings(SERPER_API_KEY="test-key", HUNT_MAX_LIMIT=100)
    @patch("leads.services.requests.post")
    def test_page_two_uses_geocoded_ll_when_page_one_has_no_coordinates(self, mock_post):
        from leads.services import fetch_leads_from_serper

        page1 = [
            {"title": f"Biz {i}", "address": f"St {i}", "phoneNumber": f"+6012{i:07d}"}
            for i in range(20)
        ]
        page2 = [
            {"title": f"Biz {i}", "address": f"St {i}", "phoneNumber": f"+6013{i:07d}"}
            for i in range(20, 25)
        ]

        def _resp(places, ll=None):
            resp = Mock()
            resp.raise_for_status = Mock()
            body: dict = {"places": places}
            if ll:
                body["searchParameters"] = {"ll": ll}
            resp.json.return_value = body
            return resp

        mock_post.side_effect = [
            _resp(page1),
            _resp([], ll="@1.4927,103.7414,13z"),
            _resp(page2, ll="@1.4927,103.7414,13z"),
        ]

        result = fetch_leads_from_serper(
            "Johor Bahru",
            "",
            num=40,
            shop_keyword="klinik",
            state="Johor",
            country="Malaysia",
        )

        self.assertEqual(mock_post.call_count, 3)
        second_payload = mock_post.call_args_list[2][1]["json"]
        self.assertEqual(second_payload["page"], 2)
        self.assertEqual(second_payload["ll"], "@1.4927,103.7414,13z")
        self.assertEqual(result.places_seen, 25)
        self.assertEqual(result.created, 25)

    @override_settings(SERPER_API_KEY="test-key", HUNT_MAX_LIMIT=100)
    @patch("leads.services.requests.post")
    def test_single_page_when_limit_is_20(self, mock_post):
        from leads.services import fetch_leads_from_serper

        places = [
            {"title": f"Solo {i}", "address": f"Road {i}", "phoneNumber": f"+6014{i:07d}"}
            for i in range(15)
        ]
        resp = Mock()
        resp.raise_for_status = Mock()
        resp.json.return_value = {"places": places}
        mock_post.return_value = resp

        result = fetch_leads_from_serper(
            "Penang",
            "",
            num=20,
            shop_keyword="gym",
            state="Penang",
            country="Malaysia",
        )

        self.assertEqual(mock_post.call_count, 1)
        self.assertEqual(result.places_seen, 15)
        self.assertEqual(result.created, 15)


class SerperExcludeKeywordTests(TestCase):
    def test_build_search_q_appends_exclude_suffix(self):
        from leads.services import _build_search_q

        q = _build_search_q(
            "Iskandar Puteri",
            "",
            shop_keyword="klinik",
            state="Johor",
            country="Malaysia",
            exclude_keywords=["dental", "24 jam"],
        )
        self.assertIn("klinik Iskandar Puteri Johor Malaysia", q)
        self.assertIn("-dental", q)
        self.assertIn('-"24 jam"', q)

    @override_settings(SERPER_API_KEY="test-key", HUNT_MAX_LIMIT=100)
    @patch("leads.services.requests.post")
    def test_fetch_skips_excluded_places(self, mock_post):
        from leads.services import fetch_leads_from_serper

        places = [
            {"title": "Alpha Klinik", "address": "St 1", "phoneNumber": "+60121111111"},
            {"title": "Beta Dental Clinic", "address": "St 2", "phoneNumber": "+60122222222"},
        ]
        resp = Mock()
        resp.raise_for_status = Mock()
        resp.json.return_value = {"places": places}
        mock_post.return_value = resp

        result = fetch_leads_from_serper(
            "Johor Bahru",
            "",
            num=20,
            shop_keyword="klinik",
            state="Johor",
            country="Malaysia",
            exclude_keywords=["dental"],
        )

        payload = mock_post.call_args[1]["json"]
        self.assertIn("-dental", payload["q"])
        self.assertEqual(result.skipped_excluded, 1)
        self.assertEqual(result.places_seen, 1)
        self.assertEqual(result.created, 1)
        self.assertEqual(Lead.objects.get().name, "Alpha Klinik")


class BackupExportTests(TestCase):
    def test_build_backup_workbook_filters_selected_leads(self):
        from io import BytesIO

        from openpyxl import load_workbook

        from leads.backup import build_backup_workbook

        group = LeadGroup.objects.create(name="Quality", sort_order=10)
        keep = Lead.objects.create(name="Keep Me", address="1 Road", group=group)
        Lead.objects.create(name="Drop Me", address="2 Road", group=group)
        ChatMessage.objects.create(lead=keep, body="Hi", is_outbound=True)
        ChatMessage.objects.create(
            lead=Lead.objects.get(name="Drop Me"),
            body="Bye",
            is_outbound=False,
        )

        wb = build_backup_workbook(lead_ids=[keep.pk])
        buf = BytesIO()
        wb.save(buf)
        loaded = load_workbook(BytesIO(buf.getvalue()), read_only=True, data_only=True)

        lead_rows = list(loaded["Leads"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(lead_rows), 1)
        self.assertEqual(lead_rows[0][1], "Keep Me")

        chat_rows = list(loaded["ChatMessages"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(chat_rows), 1)
        self.assertEqual(chat_rows[0][0], keep.pk)

        group_rows = list(loaded["Groups"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(group_rows), 1)
        self.assertEqual(group_rows[0][0], "Quality")

    def test_export_full_backup_view_accepts_ids_query(self):
        keep = Lead.objects.create(name="Export Me", address="9 Road")
        Lead.objects.create(name="Other", address="8 Road")
        client = Client()
        response = client.get(reverse("export_full_backup"), {"ids": str(keep.pk)})
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        self.assertIn(f"clinic_crm_backup_1_leads_", response["Content-Disposition"])
